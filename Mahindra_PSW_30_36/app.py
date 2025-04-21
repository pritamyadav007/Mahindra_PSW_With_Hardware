from flask import Flask, render_template, request, jsonify, redirect, url_for
import pyvisa
import os
import json
import serial
import threading
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import requests
import atexit
import serial.tools.list_ports
import webview

app = Flask(__name__)
rm = pyvisa.ResourceManager('@py')


arduino_serial = None
psw_device = None
device = None
stop_flag = False
settings_file = "settings.json"
latest_cycle_data = []
latest_log_text = ""
latest_status = "Waiting for result..."
cycle_running = False
cycle_lock = threading.Lock()


# Load or create default settings
def load_settings():
    default_settings = {
        "voltage": 24,
        "cycles": 10,
        "delay": 1,
        "username": "Pawan",
        "password": "Pawan@123",
        "excel_path": "D/PSW84"
    }

    if not os.path.exists(settings_file):
        with open(settings_file, 'w') as f:
            json.dump(default_settings, f)
        return default_settings
    else:
        with open(settings_file, 'r') as f:
            settings = json.load(f)

        for key in default_settings:
            if key not in settings:
                settings[key] = default_settings[key]

        for k in ["voltage", "delay"]:
            if isinstance(settings[k], float) and settings[k].is_integer():
                settings[k] = int(settings[k])

        with open(settings_file, 'w') as f:
            json.dump(settings, f)

        return settings

@app.route("/")
def dashboard():
    return render_template("dashboard.html")

@app.route('/get_settings')
def get_settings():
    try:
        with open("settings.json", "r") as f:
            settings = json.load(f)
        return jsonify({
            "voltage": settings.get("voltage"),
            "cycles": settings.get("cycles")
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    if request.method == "POST":
        voltage = float(request.form.get("voltage", 0))
        cycles = int(request.form.get("cycles", 0))
        delay = float(request.form.get("delay", 0))

        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                settings = json.load(f)
        else:
            settings = {}

        settings["voltage"] = voltage
        settings["cycles"] = cycles
        settings["delay"] = delay

        with open(settings_file, 'w') as f:
            json.dump(settings, f)

        return redirect(url_for("settings_page"))

    else:
        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                settings = json.load(f)
        else:
            settings = {
                "voltage": 24,
                "cycles": 5,
                "delay": 1,
                "username": "pawan",
                "password": "Pawan@123",
                "excel_path": "."
            }

        return render_template("settings.html", settings=settings)

@app.route("/verify_login", methods=["POST"])
def verify_login():
    data = request.json
    entered_username = data.get("username", "").lower()
    entered_password = data.get("password", "")
    settings = load_settings()

    stored_username = settings.get("username", "").lower()
    stored_password = settings.get("password", "")

    return jsonify({"success": entered_username == stored_username and entered_password == stored_password})

@app.route("/cycle_data")
def reading_data():
    return render_template("cycle_data.html", cycle_data=latest_cycle_data)

@app.route("/connect")
def connect():
    global psw_device, device
    try:
        if device is not None:
            return jsonify({"status": "Already Connected", "message": "Device already connected."})

        available_ports = rm.list_resources()
        serial_ports = [port for port in available_ports if "ASRL" in port]

        if not serial_ports:
            return jsonify({"status": "Error", "message": "No COM ports found"})

        psw_device = rm.open_resource(serial_ports[0])
        psw_device.baud_rate = 9600
        psw_device.data_bits = 8
        psw_device.parity = pyvisa.constants.Parity.none
        psw_device.stop_bits = pyvisa.constants.StopBits.one
        psw_device.timeout = 500

        idn = psw_device.query("*IDN?")
        device = psw_device

        return jsonify({"status": "Connected", "message": f"{idn.strip()}"})
    except Exception as e:
        return jsonify({"status": "Error", "message": f"Failed to connect: {str(e)}"})

@app.route("/disconnect")
def disconnect():
    global psw_device, device
    try:
        if psw_device:
            psw_device.close()
        psw_device = None
        device = None
        return jsonify({"status": "Disconnected", "message": "Device disconnected successfully."})
    except Exception as e:
        return jsonify({"status": "Error", "message": f"Disconnection error: {str(e)}"})

@app.route("/status")
def status():
    return jsonify({"connected": device is not None})

@app.route("/start_loop", methods=["POST"])
def start_loop():
    global device, stop_flag, latest_cycle_data, latest_log_text, latest_status, cycle_running

    with cycle_lock:
        if cycle_running:
            return jsonify({"status": "Running", "message": "Cycle already running"})
        cycle_running = True

    settings = load_settings()
    voltage = settings["voltage"]
    cycles = settings["cycles"]
    delay = settings["delay"]

    if not device:
        cycle_running = False
        return jsonify({"status": "Error", "message": "Device not connected"})

    result_data = []
    stop_flag = False
    latest_log_text = ""
    latest_status = "Cycle Running..."

    def run_cycles():
        nonlocal result_data
        global stop_flag, latest_cycle_data, latest_log_text, latest_status, cycle_running

        job_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            for i in range(1, cycles + 1):
                if stop_flag:
                    break

                try:
                    device.write("OUTP OFF")
                    device.write(f"APPL {voltage},1")
                    device.write("OUTP ON")
                    time.sleep(delay)

                    measured_voltage = float(device.query("MEAS:VOLT?"))
                    measured_current = float(device.query("MEAS:CURR?"))

                    device.write("OUTP OFF")
                    time.sleep(delay)
                    device.write("OUTP ON")

                    status = "OK" if abs(measured_voltage - voltage) < 0.1 else "Not OK"
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    result_data.append({
                        "cycle": i,
                        "voltage": round(measured_voltage, 2),
                        "current": round(measured_current, 2),
                        "status": status,
                        "timestamp": timestamp
                    })

                    latest_log_text += f"[{i}/{cycles}] Cycle {i}: V={measured_voltage:.2f}V, I={measured_current:.2f}A, Status={status}, Time={timestamp}\n"


                except Exception as e:
                    error_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    latest_log_text += f"[{i}/{cycles}] Cycle {i} Error: {str(e)} at {error_time}\n"

                    result_data.append({
                        "cycle": i,
                        "voltage": 0,
                        "current": 0,
                        "status": f"Error: {str(e)}",
                        "timestamp": error_time
                    })

            device.write("OUTP OFF")
            save_to_excel(result_data, job_start_time, voltage)

            latest_cycle_data = [
                {"Cycle": entry["cycle"], "Voltage (V)": entry["voltage"], "Current (A)": entry["current"]}
                for entry in result_data
            ]

            latest_status = "Stopped by user" if stop_flag else ("Job OK" if all(entry["status"] == "OK" for entry in result_data) else "Job Not OK")

        except Exception as err:
            latest_log_text += f"Critical error during job: {str(err)}\n"

        finally:
            cycle_running = False

    threading.Thread(target=run_cycles).start()
    return jsonify({"status": "Started", "message": "Cycle loop started"})


@app.route("/stop_loop", methods=["POST"])
def stop_loop():
    global stop_flag
    stop_flag = True
    return jsonify({"status": "Stopped", "message": "Cycle loop stopped"})

@app.route("/cycle_status")
def cycle_status():
    return jsonify({
        "logs": latest_log_text,
        "status": latest_status
    })

def save_to_excel(data, start_time, voltage):
    settings = load_settings()
    user_path = settings.get("excel_path", "").strip()
    save_dir = "."

    # Check and use user-defined path
    if user_path:
        try:
            os.makedirs(user_path, exist_ok=True)
            save_dir = user_path
        except Exception as e:
            print(f"[WARNING] Cannot use path '{user_path}', saving to current directory instead. Error: {e}")

    current_date = datetime.now().strftime("%Y-%m-%d")
    logs_excel_path = os.path.join(save_dir, f"logs_{current_date}.xlsx")
    results_excel_path = os.path.join(save_dir, f"Cycle_Result_{current_date}.xlsx")

    # Save cycle logs
    if not os.path.exists(logs_excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Log Start Time", start_time])
        ws.append(["Cycle", "Voltage", "Current", "Status", "Time"])
    else:
        wb = load_workbook(logs_excel_path)
        ws = wb.active
        ws.append([])  # Blank line for separation
        ws.append(["Log Start Time", start_time])
        ws.append(["Cycle", "Voltage", "Current", "Status", "Time"])

    for entry in data:
        ws.append([entry["cycle"], entry["voltage"], entry["current"], entry["status"], entry["timestamp"]])
    wb.save(logs_excel_path)

    # Save job summary result
    if not os.path.exists(results_excel_path):
        result_wb = Workbook()
        result_ws = result_wb.active
        result_ws.append(["Job Start", "Voltage", "Total Cycles", "Status"])
    else:
        result_wb = load_workbook(results_excel_path)
        result_ws = result_wb.active

    all_ok = all(entry["status"] == "OK" for entry in data)
    result_ws.append([start_time, voltage, len(data), "OK" if all_ok else "Not OK"])
    result_wb.save(results_excel_path)

    print(f"[INFO] Excel files saved to: {save_dir}")

# Arduino listener thread for button press
def find_arduino_port(exclude_ports=None):
    if exclude_ports is None:
        exclude_ports = []

    arduino_vids = ["2341", "1A86", "10C4"]  # CH340, CP2102, etc.
    ports = serial.tools.list_ports.comports()

    for port in ports:
        if port.device in exclude_ports:
            continue

        vid = format(port.vid, "04X") if port.vid else ""
        desc = port.description.lower()

        if vid in arduino_vids or "arduino" in desc or "ch340" in desc:
            return port.device

    return None

def listen_for_button_signals(start_url="http://127.0.0.1:5001/start_loop", stop_url="http://127.0.0.1:5001/stop_loop"):
    global device, arduino_serial
    exclude = []
    if device:
        # Convert PyVISA resource name (like "ASRL4::INSTR") to COM port (e.g. COM4)
        exclude.append(device.resource_name.replace("ASRL", "COM").replace("::INSTR", ""))

    port = find_arduino_port(exclude_ports=exclude)
    if not port:
        print("[ERROR] No Arduino found.")
        return

    ser = None  # Initialize ser outside the try block
    try:
        print(f"[INFO] Connecting to Arduino on {port}...")
        ser = serial.Serial(port, 9600, timeout=1)
        arduino_serial = ser

        while True:
            try:
                line = ser.readline().decode().strip().lower()
                if line:
                    print(f"[BUTTON] Received signal: {line}")
                    if line == "start":
                        try:
                            response = requests.post(start_url)
                            print(f"[START] {response.status_code} - {response.text}")
                        except Exception as e:
                            print(f"[ERROR] Failed to send start POST: {e}")
                    elif line == "stop":
                        try:
                            response = requests.post(stop_url)
                            print(f"[STOP] {response.status_code} - {response.text}")
                        except Exception as e:
                            print(f"[ERROR] Failed to send stop POST: {e}")
                time.sleep(0.1)

            except serial.SerialException as e:
                print(f"[SERIAL ERROR] Arduino disconnected: {e}")
                break  # Exit loop to try reconnecting

    except serial.SerialException as e:
        print(f"[ERROR] Could not open Arduino serial port {port}: {e}")
    finally:
        if ser and ser.is_open:
            print("[INFO] Closing Arduino serial port.")
            ser.close()
            arduino_serial = None
            del ser  # Explicitly delete the serial port object
            time.sleep(2)  # Try a 2-second delay

# Start Arduino listener in a background thread
def run_in_thread():
    thread = threading.Thread(target=listen_for_button_signals)
    thread.daemon = True
    thread.start()



def cleanup():
    global psw_device, arduino_serial
    if psw_device:
        try: psw_device.close()
        except: pass
    if arduino_serial and arduino_serial.is_open:
        try: arduino_serial.close()
        except: pass

atexit.register(cleanup)


if __name__ == "__main__":
    # Start Arduino button listener in the background.
    run_in_thread()
    
    # Function to run Flask in a separate thread.
    def run_flask():
        # Disable the reloader to prevent signal handling in a non-main thread.
        app.run(debug=True, use_reloader=False, port=5001)

    
    # Start Flask in a daemon thread.
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    # Optionally sleep for a couple seconds to allow the server to start.
    time.sleep(2)
    
    # Launch the PyWebView window pointing to the Flask app.
    webview.create_window("Mahindra PSW Controller", "http://127.0.0.1:5001", width=1200, height=800)
    
    # This call starts the webview event loop.
    webview.start()